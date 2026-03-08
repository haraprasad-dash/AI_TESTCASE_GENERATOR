"""
Export Service - Convert generated content to various formats.
"""
import json
import re
from pathlib import Path
from typing import Optional, Dict, Any, List
from dataclasses import dataclass

import pandas as pd


@dataclass
class ExportResult:
    """Result of export operation."""
    file_path: str
    format: str
    size_bytes: int


class ExportService:
    """Service for exporting test artifacts."""
    
    def __init__(self, output_dir: str = "./outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    async def export(
        self,
        test_plan: Optional[str],
        test_cases: Optional[str],
        format: str,
        filename_prefix: str = "testgen"
    ) -> ExportResult:
        """
        Export test artifacts to specified format.
        
        Args:
            test_plan: Test plan markdown content
            test_cases: Test cases markdown content
            format: Export format (markdown, pdf, excel, json, gherkin)
            filename_prefix: Prefix for output filename
            
        Returns:
            ExportResult with file path and metadata
        """
        if format == "markdown":
            return await self._export_markdown(test_plan, test_cases, filename_prefix)
        elif format == "pdf":
            return await self._export_pdf(test_plan, test_cases, filename_prefix)
        elif format == "excel":
            return await self._export_excel(test_cases, filename_prefix)
        elif format == "json":
            return await self._export_json(test_plan, test_cases, filename_prefix)
        elif format == "gherkin":
            return await self._export_gherkin(test_cases, filename_prefix)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def _export_markdown(
        self,
        test_plan: Optional[str],
        test_cases: Optional[str],
        filename_prefix: str
    ) -> ExportResult:
        """Export as combined Markdown file."""
        parts = []
        
        if test_plan:
            parts.append(test_plan)
            parts.append("\n\n---\n\n")
        
        if test_cases:
            parts.append(test_cases)
        
        content = "\n".join(parts)
        
        file_path = self.output_dir / f"{filename_prefix}.md"
        file_path.write_text(content, encoding='utf-8')
        
        return ExportResult(
            file_path=str(file_path),
            format="markdown",
            size_bytes=file_path.stat().st_size
        )
    
    async def _export_pdf(
        self,
        test_plan: Optional[str],
        test_cases: Optional[str],
        filename_prefix: str
    ) -> ExportResult:
        """Export as PDF file using fpdf2."""
        from fpdf import FPDF

        class PDF(FPDF):
            def header(self):
                self.set_font('Helvetica', 'B', 9)
                self.set_text_color(120, 120, 120)
                self.cell(0, 8, 'TestGen AI - Generated Test Artifacts', align='R')
                self.ln(4)

            def footer(self):
                self.set_y(-12)
                self.set_font('Helvetica', '', 8)
                self.set_text_color(160, 160, 160)
                self.cell(0, 8, f'Page {self.page_no()}', align='C')

        def safe(text: str) -> str:
            """Strip characters not supported by core PDF fonts."""
            return text.encode('latin-1', errors='replace').decode('latin-1')

        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=18)
        pdf.add_page()
        pdf.set_left_margin(15)
        pdf.set_right_margin(15)

        parts = []
        if test_plan:
            parts.append(test_plan)
        if test_cases:
            parts.append(test_cases)
        content = "\n\n---\n\n".join(parts)

        for line in content.split('\n'):
            stripped = line.strip()
            # H1
            if stripped.startswith('# ') and not stripped.startswith('## '):
                pdf.ln(4)
                pdf.set_font('Helvetica', 'B', 16)
                pdf.set_text_color(30, 30, 30)
                pdf.multi_cell(0, 9, safe(stripped[2:]))
                pdf.ln(2)
            # H2
            elif stripped.startswith('## ') and not stripped.startswith('### '):
                pdf.ln(3)
                pdf.set_font('Helvetica', 'B', 13)
                pdf.set_text_color(50, 80, 160)
                pdf.multi_cell(0, 8, safe(stripped[3:]))
                pdf.set_draw_color(180, 200, 240)
                pdf.line(15, pdf.get_y(), 195, pdf.get_y())
                pdf.ln(2)
            # H3
            elif stripped.startswith('### '):
                pdf.ln(2)
                pdf.set_font('Helvetica', 'B', 11)
                pdf.set_text_color(60, 60, 60)
                pdf.multi_cell(0, 7, safe(stripped[4:]))
                pdf.ln(1)
            # H4
            elif stripped.startswith('#### '):
                pdf.ln(1)
                pdf.set_font('Helvetica', 'BI', 10)
                pdf.set_text_color(80, 80, 80)
                pdf.multi_cell(0, 6, safe(stripped[5:]))
            # Horizontal rule
            elif re.match(r'^[-*]{3,}$', stripped):
                pdf.ln(2)
                pdf.set_draw_color(200, 200, 200)
                pdf.line(15, pdf.get_y(), 195, pdf.get_y())
                pdf.ln(3)
            # Table separator row — skip
            elif re.match(r'^\|[-:\s|]+\|$', stripped):
                continue
            # Table data row
            elif stripped.startswith('|') and stripped.endswith('|'):
                cells = [c.strip() for c in stripped.split('|')[1:-1]]
                if not cells:
                    continue
                col_w = min(int(175 / len(cells)), 55)
                pdf.set_font('Helvetica', '', 8)
                pdf.set_text_color(30, 30, 30)
                # Detect if first row (no separator seen yet means it's a header)
                for idx, cell in enumerate(cells):
                    clean = re.sub(r'\*\*(.*?)\*\*', r'\1', cell)
                    clean = re.sub(r'\*(.*?)\*', r'\1', clean)
                    clean = safe(clean[:60])
                    pdf.cell(col_w, 6, clean, border=1)
                pdf.ln()
            # Unordered list
            elif re.match(r'^[-*+] ', stripped):
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(30, 30, 30)
                text = re.sub(r'\*\*(.*?)\*\*', r'\1', stripped[2:])
                text = re.sub(r'\*(.*?)\*', r'\1', text)
                indent = (len(line) - len(line.lstrip())) // 2
                pdf.set_x(15 + indent * 4)
                pdf.multi_cell(0, 6, safe(f'\u2022 {text}'))
            # Ordered list
            elif re.match(r'^\d+[\.\)] ', stripped):
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(30, 30, 30)
                text = re.sub(r'\*\*(.*?)\*\*', r'\1', stripped)
                text = re.sub(r'\*(.*?)\*', r'\1', text)
                pdf.multi_cell(0, 6, safe(text))
            # Blockquote
            elif stripped.startswith('> '):
                pdf.set_font('Helvetica', 'I', 10)
                pdf.set_text_color(90, 90, 90)
                pdf.set_x(20)
                pdf.multi_cell(0, 6, safe(stripped[2:]))
            # Empty line
            elif stripped == '':
                pdf.ln(2)
            # Normal paragraph
            else:
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(30, 30, 30)
                clean = re.sub(r'\*\*(.*?)\*\*', r'\1', stripped)
                clean = re.sub(r'\*(.*?)\*', r'\1', clean)
                clean = re.sub(r'`(.*?)`', r'\1', clean)
                pdf.multi_cell(0, 6, safe(clean))

        file_path = self.output_dir / f"{filename_prefix}.pdf"
        pdf.output(str(file_path))

        return ExportResult(
            file_path=str(file_path),
            format="pdf",
            size_bytes=file_path.stat().st_size
        )
    
    async def _export_excel(
        self,
        test_cases: Optional[str],
        filename_prefix: str
    ) -> ExportResult:
        """Export test cases as Excel file."""
        if not test_cases:
            raise ExportError("No test cases to export")
        
        # Parse test cases from markdown tables
        parsed_cases = self._parse_test_cases(test_cases)
        
        if not parsed_cases:
            raise ExportError("Could not parse test cases")
        
        file_path = self.output_dir / f"{filename_prefix}.xlsx"
        
        # Create DataFrame
        df = pd.DataFrame(parsed_cases)
        
        # Export to Excel with formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Test Cases', index=False)

            worksheet = writer.sheets['Test Cases']

            # Style header row
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            thin = Side(style='thin', color='CBD5E1')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, cell in enumerate(worksheet[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.border = border

            # Style data rows with alternating shading
            light_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
            data_font = Font(size=10)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                fill = light_fill if row_idx % 2 == 0 else None
                for cell in row:
                    if fill:
                        cell.fill = fill
                    cell.font = data_font
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border

            # Auto-adjust column widths (capped)
            for column in worksheet.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in column if cell.value is not None),
                    default=8
                )
                col_letter = column[0].column_letter
                worksheet.column_dimensions[col_letter].width = min(max_length + 4, 55)

            # Freeze header row
            worksheet.freeze_panes = 'A2'
        
        return ExportResult(
            file_path=str(file_path),
            format="excel",
            size_bytes=file_path.stat().st_size
        )
    
    async def _export_json(
        self,
        test_plan: Optional[str],
        test_cases: Optional[str],
        filename_prefix: str
    ) -> ExportResult:
        """Export as structured JSON."""
        data = {
            "test_plan": test_plan,
            "test_cases": self._parse_test_cases(test_cases) if test_cases else [],
            "metadata": {
                "format_version": "1.0",
                "export_date": str(pd.Timestamp.now())
            }
        }
        
        file_path = self.output_dir / f"{filename_prefix}.json"
        file_path.write_text(
            json.dumps(data, indent=2, ensure_ascii=False),
            encoding='utf-8'
        )
        
        return ExportResult(
            file_path=str(file_path),
            format="json",
            size_bytes=file_path.stat().st_size
        )
    
    async def _export_gherkin(
        self,
        test_cases: Optional[str],
        filename_prefix: str
    ) -> ExportResult:
        """Export test cases as Gherkin .feature file."""
        if not test_cases:
            raise ExportError("No test cases to export")
        
        gherkin_content = self._convert_to_gherkin(test_cases)
        
        file_path = self.output_dir / f"{filename_prefix}.feature"
        file_path.write_text(gherkin_content, encoding='utf-8')
        
        return ExportResult(
            file_path=str(file_path),
            format="gherkin",
            size_bytes=file_path.stat().st_size
        )
    
    def _strip_markdown(self, text: str) -> str:
        """Remove markdown inline formatting from a plain-text string."""
        text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)   # **bold**
        text = re.sub(r'\*(.*?)\*', r'\1', text)         # *italic*
        text = re.sub(r'__(.*?)__', r'\1', text)          # __bold__
        text = re.sub(r'_(.*?)_', r'\1', text)            # _italic_
        text = re.sub(r'`(.*?)`', r'\1', text)            # `code`
        text = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', text)  # [link](url)
        return text.strip()

    def _find_col(self, case: Dict[str, str], *keywords) -> str:
        """Return the first non-empty cell value whose column header contains any keyword (case-insensitive)."""
        for key, val in case.items():
            if not (val and val.strip()):
                continue
            key_lower = key.lower()
            for kw in keywords:
                if kw.lower() in key_lower:
                    return val.strip()
        return ''

    def _build_scenario_title(self, case: Dict[str, str], index: int) -> str:
        """Compose a meaningful Gherkin scenario title from available columns."""
        description = self._find_col(case, 'description')
        feature     = self._find_col(case, 'feature', 'capability', 'module')
        test_id     = self._find_col(case, 'test id', 'test_id')

        if description and feature:
            return f"[{feature}] {description}"
        if description:
            return description
        if feature and test_id:
            return f"[{feature}] {test_id}"
        if test_id:
            return f"Test {test_id}"
        # Last resort: use a snippet of the BDD steps as the title
        steps_snippet = self._find_col(case, 'bdd', 'step', 'given')
        if steps_snippet:
            # Take up to 80 chars, stopping at the first keyword boundary
            short = re.split(r'\b(?:When|Then|And|But)\b', steps_snippet)[0].strip()
            prefix = re.sub(r'^Given\s+', '', short, flags=re.IGNORECASE)
            return prefix[:80].rstrip(',;')
        return f'Test Case {index}'

    def _parse_test_cases(self, markdown_content: str) -> List[Dict[str, str]]:
        """Parse test cases from markdown tables, stripping all markdown formatting."""
        cases = []

        # Match header row | sep row | data rows
        table_pattern = r'\|([^\n]+)\|\n\|[-:\|\s]+\|\n((?:\|[^\n]+\|\n?)+)'
        matches = re.findall(table_pattern, markdown_content)

        for header_line, rows in matches:
            headers = [self._strip_markdown(h.strip()) for h in header_line.split('|') if h.strip()]

            for row in rows.strip().split('\n'):
                cells = [self._strip_markdown(c.strip()) for c in row.split('|') if c.strip()]
                if len(cells) >= len(headers):
                    case = dict(zip(headers, cells))
                    cases.append(case)

        return cases

    def _convert_to_gherkin(self, test_cases: str) -> str:
        """Convert test cases to Gherkin .feature format."""
        lines = [
            "Feature: Generated Test Cases",
            "",
            "  Generated from TestGen AI Agent",
            ""
        ]

        parsed = self._parse_test_cases(test_cases)
        GWT = {'given', 'when', 'then', 'and', 'but'}

        for i, case in enumerate(parsed, 1):
            title        = self._build_scenario_title(case, i)
            steps_text   = self._find_col(case, 'bdd', 'step', 'given', 'scenario')
            precondition = self._find_col(case, 'precondition', 'prerequisite', 'pre-condition')
            expected     = self._find_col(case, 'expected', 'result', 'outcome')

            lines.append(f"  Scenario: {title}")

            # ── Strategy: split on GWT keywords already present in the cell ──
            # The AI typically writes "Given X When Y Then Z" all in one table cell.
            parts = re.split(r'\b(Given|When|Then|And|But)\b', steps_text)
            has_gwt_keywords = any(p in ('Given', 'When', 'Then', 'And', 'But') for p in parts)

            if has_gwt_keywords:
                keyword = None
                pending_parts: list = []
                for part in parts:
                    part = part.strip()
                    if not part:
                        continue
                    if part in ('Given', 'When', 'Then', 'And', 'But'):
                        # Flush accumulated text under previous keyword
                        if keyword and pending_parts:
                            lines.append(f"    {keyword} {' '.join(pending_parts)}")
                        keyword = part
                        pending_parts = []
                    else:
                        pending_parts.append(part)
                # Flush final keyword
                if keyword and pending_parts:
                    lines.append(f"    {keyword} {' '.join(pending_parts)}")
            else:
                # No GWT keywords found — build from precondition + numbered steps + expected
                if precondition:
                    lines.append(f"    Given {precondition}")

                if steps_text:
                    step_parts = [s.strip() for s in re.split(r'\d+[\.\)]\s+|[;\n]', steps_text) if s.strip()]
                    for j, step in enumerate(step_parts):
                        if not precondition and j == 0:
                            lines.append(f"    Given {step}")
                        elif j == len(step_parts) - 1:
                            lines.append(f"    Then {step}")
                        else:
                            lines.append(f"    When {step}" if j == 1 else f"    And {step}")
                elif expected:
                    lines.append(f"    Then {expected}")

            lines.append("")

        return "\n".join(lines)


class ExportError(Exception):
    """Export operation failed."""
    pass
